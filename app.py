from __future__ import annotations

import copy
import json
import hashlib
import threading
import time
from pathlib import Path

import os
import httpx

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import Response as RawResponse
from collections import Counter
from html import escape

from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

from converter_engine import InMemoryJobStore, LowRelevanceError, QCVWebEngine, make_temp_workspace, resolve_api_key, _build_output_base_name, choose_model_name, configure_gemini, call_llm_json
import cv_engine as _core
import auth as _auth

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(APP_DIR)))
TEMPLATES_DIR = APP_DIR / "templates"
STORE_DIR = DATA_DIR / "_store"
JD_STORE_DIR = DATA_DIR / "_jd_store"
USAGE_LOG = DATA_DIR / "usage_log.jsonl"

TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
STORE_DIR.mkdir(parents=True, exist_ok=True)
JD_STORE_DIR.mkdir(parents=True, exist_ok=True)
_auth.init(DATA_DIR)

app = FastAPI(title="Q-CV Web Converter")
app.mount("/images", StaticFiles(directory=APP_DIR / "images"), name="images")
jobs = InMemoryJobStore()
_SERVER_START = time.time()
_stats_cache: dict = {}
_stats_cache_ts: float = 0.0
_JOB_SEMAPHORE = threading.Semaphore(10)  # max 10 concurrent LLM jobs
_STORE_LOCK = threading.RLock()  # serialize store file writes (reentrant for nested calls)
_batch_cancel_flags: dict[str, bool] = {}  # batch_id -> cancelled flag

# ── Embedding cache for fast match ──
import numpy as np
EMBED_CACHE_PATH = DATA_DIR / "_cache" / "embeddings.npz"
(DATA_DIR / "_cache").mkdir(parents=True, exist_ok=True)
_embed_ids: list[str] = []      # store_ids in order
_embed_matrix: np.ndarray | None = None  # (N, dim) float32
_EMBED_LOCK = threading.Lock()

def _load_embed_cache():
    global _embed_ids, _embed_matrix
    if EMBED_CACHE_PATH.exists():
        data = np.load(EMBED_CACHE_PATH, allow_pickle=True)
        _embed_ids = data["ids"].tolist()
        _embed_matrix = data["vecs"].astype(np.float32)

def _save_embed_cache():
    if _embed_matrix is not None and _embed_ids:
        np.savez_compressed(EMBED_CACHE_PATH,
                            ids=np.array(_embed_ids, dtype=object),
                            vecs=_embed_matrix)

def _add_embedding(store_id: str, vec: list[float]):
    global _embed_ids, _embed_matrix
    with _EMBED_LOCK:
        arr = np.array(vec, dtype=np.float32).reshape(1, -1)
        if store_id in _embed_ids:
            idx = _embed_ids.index(store_id)
            _embed_matrix[idx] = arr[0]
        else:
            _embed_ids.append(store_id)
            if _embed_matrix is None:
                _embed_matrix = arr
            else:
                _embed_matrix = np.vstack([_embed_matrix, arr])
        _save_embed_cache()

def _cv_text_for_embedding(cv_json: dict) -> str:
    """Build compact text representation of CV for embedding."""
    basics = cv_json.get("basics", {})
    parts = [basics.get("current_title", "")]
    # Skills
    skills = cv_json.get("skills", {})
    for cat, items in skills.items():
        if isinstance(items, list):
            parts.append(", ".join(items))
        elif isinstance(items, str):
            parts.append(items)
    # Summary
    parts.append(cv_json.get("summary", ""))
    # Experience highlights + environment (top 3 roles)
    for exp in cv_json.get("experience", [])[:3]:
        parts.append(exp.get("role", ""))
        hl = exp.get("highlights", [])
        if isinstance(hl, list):
            for h in hl[:3]:
                if isinstance(h, str):
                    parts.append(h)
        env = exp.get("environment", "")
        if isinstance(env, str):
            parts.append(env)
    return " ".join(str(p) for p in parts if p)[:2000]  # cap at 2000 chars

def _compute_embedding(text: str) -> list[float] | None:
    """Call Gemini embedding API."""
    try:
        from google import genai
        config = _core.load_config()
        api_key = resolve_api_key(APP_DIR, config)
        client = genai.Client(api_key=api_key)
        result = client.models.embed_content(
            model="gemini-embedding-001",
            contents=text,
        )
        return result.embeddings[0].values
    except Exception as e:
        print(f"⚠️ Embedding failed: {e}")
        return None

def _cosine_similarity_batch(jd_vec: np.ndarray, matrix: np.ndarray) -> np.ndarray:
    """Compute cosine similarity between JD vector and all CV vectors."""
    jd_norm = jd_vec / (np.linalg.norm(jd_vec) + 1e-10)
    norms = np.linalg.norm(matrix, axis=1, keepdims=True) + 1e-10
    normed = matrix / norms
    return (normed @ jd_norm.T).flatten()

_load_embed_cache()

# ── Keyword matching for fast scoring ──
_KW_STOP_WORDS = {
    "the", "a", "an", "and", "or", "of", "to", "in", "for", "with", "on", "at", "by",
    "is", "are", "was", "were", "be", "been", "being", "have", "has", "had", "do", "does",
    "will", "would", "could", "should", "may", "might", "shall", "can", "must",
    "not", "no", "but", "if", "then", "than", "that", "this", "these", "those",
    "it", "its", "we", "our", "you", "your", "they", "their", "he", "she",
    "as", "from", "about", "into", "through", "during", "before", "after",
    "above", "below", "between", "under", "over", "such", "each", "all", "any",
    "both", "more", "most", "other", "some", "only", "very", "also", "just",
    "experience", "required", "requirements", "responsibilities", "ability",
    "work", "working", "team", "role", "position", "candidate", "including",
    "strong", "knowledge", "skills", "years", "etc", "using", "used",
    "provide", "ensure", "manage", "develop", "support", "include",
    "based", "related", "well", "new", "within", "across",
}

def _extract_keywords(text: str) -> set[str]:
    """Extract meaningful keywords from text (lowercase, no stop words, min 2 chars)."""
    tokens = re.findall(r'[a-zA-Z0-9#+./-]+', text.lower())
    # Keep multi-word tech terms by also extracting bigrams
    keywords = set()
    for t in tokens:
        if len(t) >= 2 and t not in _KW_STOP_WORDS:
            keywords.add(t)
    return keywords

def _cv_keywords(cv_json: dict) -> set[str]:
    """Extract keywords from structured CV JSON."""
    parts = []
    basics = cv_json.get("basics", {})
    parts.append(basics.get("current_title", ""))
    # Skills — most important
    skills = cv_json.get("skills", {})
    for cat, items in skills.items():
        if isinstance(items, list):
            parts.extend(str(i) for i in items)
        elif isinstance(items, str):
            parts.append(items)
    # Summary
    parts.append(str(cv_json.get("summary", "")))
    # Experience environment (tech stacks)
    for exp in cv_json.get("experience", [])[:5]:
        env = exp.get("environment", "")
        if isinstance(env, str):
            parts.append(env)
        parts.append(str(exp.get("role", "")))
    return _extract_keywords(" ".join(parts))

# Pre-compute CV keywords cache
_cv_keywords_cache: dict[str, set[str]] = {}
_CV_KW_LOCK = threading.Lock()

def _get_cv_keywords(store_id: str) -> set[str]:
    """Get cached keywords for a CV, computing if needed."""
    with _CV_KW_LOCK:
        if store_id in _cv_keywords_cache:
            return _cv_keywords_cache[store_id]
    # Compute outside lock
    p = STORE_DIR / f"{store_id}.json"
    if not p.exists():
        return set()
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        cv = {k: v for k, v in data.items() if not k.startswith("_")}
        kw = _cv_keywords(cv)
        with _CV_KW_LOCK:
            _cv_keywords_cache[store_id] = kw
        return kw
    except Exception:
        return set()

def _keyword_match_score(jd_keywords: set[str], cv_keywords: set[str]) -> int:
    """Compute keyword overlap percentage: matched JD keywords / total JD keywords."""
    if not jd_keywords:
        return 0
    matched = jd_keywords & cv_keywords
    return int(len(matched) / len(jd_keywords) * 100)


import re
_STORE_ID_RE = re.compile(r'^[a-fA-F0-9]+$')

def _validate_store_id(store_id: str) -> None:
    """Reject path traversal and invalid store IDs."""
    if not store_id or not _STORE_ID_RE.match(store_id):
        raise HTTPException(status_code=400, detail="Invalid store ID")

# Background cleanup: remove finished jobs and their tmp dirs every 10 minutes
_JOB_MAX_AGE_SEC = 3600  # 1 hour

def _cleanup_loop():
    import shutil
    while True:
        time.sleep(600)
        try:
            removed = jobs.cleanup_old(_JOB_MAX_AGE_SEC)
            # Clean orphaned tmp dirs older than max age
            tmp_root = Path(os.environ.get("TMPDIR", "/tmp"))
            cutoff = time.time() - _JOB_MAX_AGE_SEC
            for d in tmp_root.glob("qcv_web_*"):
                if d.is_dir() and d.stat().st_mtime < cutoff:
                    shutil.rmtree(d, ignore_errors=True)
        except Exception:
            pass

threading.Thread(target=_cleanup_loop, daemon=True).start()


def _backfill_search_text():
    """One-time migration: add search_text to store entries that lack it."""
    for p in STORE_DIR.glob("*.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            meta = data.get("_meta", {})
            if "search_text" in meta:
                continue
            basics = data.get("basics", {})
            exp = data.get("experience", [])
            skills_text = json.dumps(data.get("skills", {}), ensure_ascii=False).lower()
            meta["search_text"] = " ".join([
                basics.get("name", ""),
                basics.get("current_title", ""),
                exp[0].get("company_name", "") if exp else "",
                meta.get("source_filename", ""),
                meta.get("comments", ""),
                skills_text,
            ]).lower()
            data["_meta"] = meta
            p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            continue

_backfill_search_text()
# Cache will be initialized lazily on first _list_store() call


def append_usage(event: dict) -> None:
    event = dict(event)
    event.setdefault("ts", time.strftime("%Y-%m-%dT%H:%M:%S"))
    with USAGE_LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")

def _read_usage_events() -> list[dict]:
    if not USAGE_LOG.exists():
        return []

    events: list[dict] = []
    for line in USAGE_LOG.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            item = json.loads(line)
        except Exception:
            continue
        if isinstance(item, dict):
            events.append(item)
    return events


_GEMINI_BASE = "https://generativelanguage.googleapis.com"
_PROXY_HOP_HEADERS = {"host", "content-length", "transfer-encoding", "connection"}


async def _proxy_to_gemini(request: Request, path: str) -> RawResponse:
    target_url = f"{_GEMINI_BASE}/{path}"
    body = await request.body()
    headers = {k: v for k, v in request.headers.items()
               if k.lower() not in _PROXY_HOP_HEADERS}
    params = dict(request.query_params)
    try:
        async with httpx.AsyncClient(timeout=300.0) as client:
            resp = await client.request(
                method=request.method,
                url=target_url,
                headers=headers,
                params=params,
                content=body,
            )
        resp_headers = {k: v for k, v in resp.headers.items()
                        if k.lower() not in {"transfer-encoding", "content-encoding"}}
        return RawResponse(
            content=resp.content,
            status_code=resp.status_code,
            headers=resp_headers,
            media_type=resp.headers.get("content-type"),
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Proxy error: {e}")


@app.api_route("/v1beta/{path:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE"])
async def gemini_proxy_v1beta(request: Request, path: str):
    """Proxy for Gemini API v1beta calls (generate content, file get/delete)."""
    return await _proxy_to_gemini(request, f"v1beta/{path}")


@app.api_route("/v1/{path:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE"])
async def gemini_proxy_v1(request: Request, path: str):
    """Proxy for Gemini API v1 (stable) calls."""
    return await _proxy_to_gemini(request, f"v1/{path}")


@app.api_route("/v1alpha/{path:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE"])
async def gemini_proxy_v1alpha(request: Request, path: str):
    """Proxy for Gemini API v1alpha (experimental) calls."""
    return await _proxy_to_gemini(request, f"v1alpha/{path}")


@app.api_route("/upload/{path:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE"])
async def gemini_proxy_upload(request: Request, path: str):
    """Proxy for Gemini Files API uploads."""
    return await _proxy_to_gemini(request, f"upload/{path}")


@app.get("/admin/usage/data")
def admin_usage_data(request: Request):
    """Return usage stats as JSON for the dashboard."""
    _auth.require_role(_auth.ADMIN)(request)
    events = _read_usage_events()

    started = [e for e in events if e.get("event") == "started"]
    done = [e for e in events if e.get("event") == "done"]
    failed = [e for e in events if e.get("event") == "failed"]

    # Durations
    durations = [e["duration_sec"] for e in done if e.get("duration_sec")]
    avg_dur = round(sum(durations) / len(durations), 1) if durations else 0

    # Unique users
    users = Counter(e.get("user") or e.get("ip", "unknown") for e in started)
    unique_users = len(users)

    # Today / this week
    today = time.strftime("%Y-%m-%d")
    week_ago = time.strftime("%Y-%m-%d", time.localtime(time.time() - 7 * 86400))
    today_jobs = sum(1 for e in started if e.get("ts", "").startswith(today))
    week_jobs = sum(1 for e in started if e.get("ts", "")[:10] >= week_ago)

    # Daily activity (last 30 days)
    day_counter: dict[str, dict] = {}
    for e in events:
        day = e.get("ts", "")[:10]
        if not day:
            continue
        if day not in day_counter:
            day_counter[day] = {"done": 0, "failed": 0, "started": 0}
        ev = e.get("event", "")
        if ev in day_counter[day]:
            day_counter[day][ev] += 1
    # Last 30 days only
    cutoff_30 = time.strftime("%Y-%m-%d", time.localtime(time.time() - 30 * 86400))
    daily = sorted(
        [(d, c) for d, c in day_counter.items() if d >= cutoff_30],
        key=lambda x: x[0],
    )

    # Hourly distribution
    hours = [0] * 24
    for e in started:
        ts = e.get("ts", "")
        if len(ts) >= 13:
            try:
                hours[int(ts[11:13])] += 1
            except (ValueError, IndexError):
                pass

    # Top users
    top_users = users.most_common(20)

    # Templates
    templates = Counter(
        (e.get("template") or "").replace(".docx", "")
        for e in started if e.get("template")
    ).most_common(10)

    # Options usage
    tailor_count = sum(1 for e in started if e.get("tailor"))
    anon_count = sum(1 for e in started if e.get("anonymize"))
    autofix_count = sum(1 for e in started if e.get("autofix"))
    total = len(started) or 1

    # Recent errors
    recent_errors = [
        {"ts": e.get("ts", ""), "user": e.get("user", ""), "file": e.get("file", ""),
         "error": e.get("error", ""), "duration": e.get("duration_sec", "")}
        for e in reversed(failed)
    ][:20]

    # Recent events (last 100)
    recent = [
        {"ts": e.get("ts", ""), "event": e.get("event", ""), "user": e.get("user", ""),
         "file": e.get("file", ""), "template": e.get("template", ""),
         "duration": e.get("duration_sec", ""), "tailor": e.get("tailor", False),
         "error": e.get("error", "")}
        for e in reversed(events)
    ][:100]

    return {
        "summary": {
            "total": len(started), "done": len(done), "failed": len(failed),
            "today": today_jobs, "week": week_jobs,
            "unique_users": unique_users, "avg_duration": avg_dur,
            "success_rate": round(len(done) / total * 100, 1),
        },
        "daily": [{"date": d, **c} for d, c in daily],
        "hours": hours,
        "top_users": [{"user": u, "count": c} for u, c in top_users],
        "templates": [{"name": n, "count": c} for n, c in templates],
        "options": {
            "tailor": tailor_count, "anonymize": anon_count,
            "autofix": autofix_count, "total": len(started),
        },
        "recent_errors": recent_errors,
        "recent": recent,
    }


@app.get("/admin/usage", response_class=HTMLResponse)
def admin_usage(request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    return HTMLResponse((Path(__file__).parent / "templates" / "admin_usage.html").read_text(encoding="utf-8"))


@app.get("/admin/prompts")
def get_prompts(request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    cfg = _core.load_config()
    prompts = {k: cfg[k] for k in cfg if k.startswith("prompt_")}
    defaults = dict(_core.DEFAULT_PROMPTS)
    return {"prompts": prompts, "defaults": defaults}


@app.put("/admin/prompts/{key}")
async def save_prompt(key: str, request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    if not key.startswith("prompt_"):
        raise HTTPException(status_code=400, detail="Invalid prompt key")
    body = await request.json()
    cfg = _core.load_config()
    cfg[key] = body["text"]
    _core.save_config(cfg)
    return {"ok": True}


@app.delete("/admin/prompts/{key}")
def reset_prompt(key: str, request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    if key not in _core.DEFAULT_PROMPTS:
        raise HTTPException(status_code=404, detail="Unknown prompt key")
    cfg = _core.load_config()
    cfg[key] = _core.DEFAULT_PROMPTS[key]
    _core.save_config(cfg)
    return {"ok": True, "text": cfg[key]}


## ── Store endpoints (Batch tab) ──────────────────────────────────────

@app.get("/store")
def list_store(request: Request, jd_id: str = ""):
    _auth.require_auth(request)
    items = _list_store()
    # Compute fast_match via embeddings if JD selected and embeddings available
    fast_scores: dict[str, int] = {}
    if jd_id and _embed_matrix is not None and len(_embed_ids) > 0:
        jd_path = JD_STORE_DIR / f"{jd_id}.json"
        if jd_path.exists():
            try:
                jd_data = json.loads(jd_path.read_text(encoding="utf-8"))
                jd_text = jd_data.get("text", "")
                if jd_text.strip():
                    # Cache JD embeddings in memory
                    if not hasattr(list_store, "_jd_embed_cache"):
                        list_store._jd_embed_cache = {}
                    if jd_id not in list_store._jd_embed_cache:
                        jd_vec = _compute_embedding(jd_text[:2000])
                        if jd_vec:
                            list_store._jd_embed_cache[jd_id] = np.array(jd_vec, dtype=np.float32)
                    if jd_id in list_store._jd_embed_cache:
                        with _EMBED_LOCK:
                            scores = _cosine_similarity_batch(list_store._jd_embed_cache[jd_id], _embed_matrix)
                            # Min-max normalize to 0-100 for visual clarity
                            smin, smax = float(scores.min()), float(scores.max())
                            spread = smax - smin if smax > smin else 1.0
                            for i, sid in enumerate(_embed_ids):
                                fast_scores[sid] = max(0, min(100, int((scores[i] - smin) / spread * 100)))
            except Exception as e:
                print(f"⚠️ Fast match error: {e}")

    # Compute keyword match scores
    key_scores: dict[str, int] = {}
    if jd_id:
        jd_path = JD_STORE_DIR / f"{jd_id}.json"
        if jd_path.exists():
            try:
                jd_data = json.loads(jd_path.read_text(encoding="utf-8"))
                jd_text = jd_data.get("text", "")
                if jd_text.strip():
                    jd_kw = _extract_keywords(jd_text)
                    raw_scores = {}
                    for m in items:
                        sid = m.get("id", "")
                        cv_kw = _get_cv_keywords(sid)
                        raw_scores[sid] = _keyword_match_score(jd_kw, cv_kw)
                    # Min-max normalize to 0-100
                    if raw_scores:
                        vals = raw_scores.values()
                        smin, smax = min(vals), max(vals)
                        spread = smax - smin if smax > smin else 1
                        for sid, v in raw_scores.items():
                            key_scores[sid] = int((v - smin) / spread * 100)
            except Exception as e:
                print(f"⚠️ Keyword match error: {e}")

    if not jd_id:
        for m in items:
            m["match_pct"] = None
            m["fast_match"] = None
            m["key_match"] = None
        return {"items": items}

    # Override match_pct with LLM score for specific JD + add fast_match + key_match
    for m in items:
        sid = m.get("id", "")
        m["fast_match"] = fast_scores.get(sid)
        m["key_match"] = key_scores.get(sid)
        p = STORE_DIR / f"{sid}.json"
        if p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                gap_sessions = data.get("_gap_sessions", {})
                if jd_id in gap_sessions:
                    m["match_pct"] = int(gap_sessions[jd_id].get("gap_analysis", {}).get("match_percentage", 0))
                else:
                    old = data.get("_gap_session")
                    if old and old.get("jd_text"):
                        old_jd_id = hashlib.sha256(old["jd_text"].strip().encode()).hexdigest()
                        if old_jd_id == jd_id:
                            m["match_pct"] = int(old.get("gap_analysis", {}).get("match_percentage", 0))
                        else:
                            m["match_pct"] = None
                    else:
                        m["match_pct"] = None
            except Exception:
                m["match_pct"] = None
        else:
            m["match_pct"] = None
    return {"items": items}


@app.post("/store/reindex_embeddings")
async def reindex_embeddings(request: Request):
    """Recompute embeddings for all CVs missing from cache. Returns progress."""
    _auth.require_auth(request)
    missing = []
    for p in STORE_DIR.glob("*.json"):
        sid = p.stem
        if sid not in _embed_ids:
            missing.append(sid)
    if not missing:
        return {"ok": True, "total": len(_embed_ids), "computed": 0, "message": "All embeddings up to date"}

    computed = 0
    errors = 0
    for sid in missing:
        try:
            data = json.loads((STORE_DIR / f"{sid}.json").read_text(encoding="utf-8"))
            cv = {k: v for k, v in data.items() if not k.startswith("_")}
            text = _cv_text_for_embedding(cv)
            if not text.strip():
                continue
            vec = _compute_embedding(text)
            if vec:
                _add_embedding(sid, vec)
                computed += 1
            else:
                errors += 1
        except Exception as e:
            print(f"⚠️ Embed error for {sid[:8]}: {e}")
            errors += 1
    return {"ok": True, "total": len(_embed_ids), "computed": computed, "errors": errors}


@app.get("/store/embedding_stats")
def embedding_stats(request: Request):
    """Return embedding cache stats."""
    _auth.require_auth(request)
    store_count = len(list(STORE_DIR.glob("*.json")))
    return {
        "store_count": store_count,
        "embedded_count": len(_embed_ids),
        "missing": store_count - len(_embed_ids),
        "dimensions": _embed_matrix.shape[1] if _embed_matrix is not None else 0,
    }


@app.get("/store/{store_id}")
def get_store_item(store_id: str, request: Request):
    """Return full stored CV including tailor session if present."""
    _auth.require_auth(request)
    _validate_store_id(store_id)
    p = STORE_DIR / f"{store_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="Not found")
    data = json.loads(p.read_text(encoding="utf-8"))
    return data


@app.delete("/store/{store_id}")
def delete_store_item(store_id: str, request: Request):
    _auth.require_auth(request)
    _validate_store_id(store_id)
    p = STORE_DIR / f"{store_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="Not found")
    p.unlink()
    _store_cache_remove(store_id)
    # Clean up caches: base_json, embeddings, keywords
    base_cache = DATA_DIR / "_cache" / "base_json" / f"{store_id}.base.json"
    if base_cache.exists():
        base_cache.unlink()
    global _embed_matrix
    with _EMBED_LOCK:
        if store_id in _embed_ids:
            idx = _embed_ids.index(store_id)
            _embed_ids.pop(idx)
            if _embed_matrix is not None and idx < len(_embed_matrix):
                _embed_matrix = np.delete(_embed_matrix, idx, axis=0)
            _save_embed_cache()
    _cv_keywords_cache.pop(store_id, None)
    return {"ok": True}


_EDITABLE_META_FIELDS = {"comments"}

@app.patch("/store/{store_id}/meta")
async def update_store_meta(store_id: str, request: Request):
    _auth.require_auth(request)
    _validate_store_id(store_id)
    body = await request.json()
    field = body.get("field", "")
    value = body.get("value", "")
    if field not in _EDITABLE_META_FIELDS:
        raise HTTPException(status_code=400, detail=f"Field not editable: {field}")
    with _STORE_LOCK:
        p = STORE_DIR / f"{store_id}.json"
        if not p.exists():
            raise HTTPException(status_code=404, detail="Not found")
        data = json.loads(p.read_text(encoding="utf-8"))
        data["_meta"][field] = value
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        cached = _store_cache_get_meta(store_id)
        if cached:
            cached[field] = value
    return {"ok": True}


@app.post("/store/batch")
async def batch_store_action(request: Request):
    _auth.require_role(_auth.ADMIN, _auth.USER)(request)
    body = await request.json()
    action = body.get("action")
    ids = body.get("ids", [])
    template_name = body.get("template_name", "")
    jd_text = body.get("jd_text", "")
    anonymize = body.get("anonymize", False)

    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")

    for sid in ids:
        _validate_store_id(sid)

    if action == "delete":
        deleted = 0
        global _embed_matrix
        for sid in ids:
            p = STORE_DIR / f"{sid}.json"
            if p.exists():
                p.unlink()
                _store_cache_remove(sid)
                # Clean up caches (same as single delete)
                base_cache = DATA_DIR / "_cache" / "base_json" / f"{sid}.base.json"
                if base_cache.exists():
                    base_cache.unlink()
                with _EMBED_LOCK:
                    if sid in _embed_ids:
                        idx = _embed_ids.index(sid)
                        _embed_ids.pop(idx)
                        if _embed_matrix is not None and idx < len(_embed_matrix):
                            _embed_matrix = np.delete(_embed_matrix, idx, axis=0)
                _cv_keywords_cache.pop(sid, None)
                deleted += 1
        if deleted:
            with _EMBED_LOCK:
                _save_embed_cache()
        return {"ok": True, "deleted": deleted}

    if action == "analyze":
        if not jd_text.strip():
            raise HTTPException(status_code=400, detail="JD text is required for analysis")
        # Filter out already analyzed with same JD
        todo_ids = []
        skipped = 0
        jd_id = hashlib.sha256(jd_text.strip().encode()).hexdigest()
        for sid in ids:
            p = STORE_DIR / f"{sid}.json"
            if p.exists():
                try:
                    store_data = json.loads(p.read_text(encoding="utf-8"))
                    # Check new format first, then old
                    gap_sessions = store_data.get("_gap_sessions", {})
                    if jd_id in gap_sessions:
                        skipped += 1
                        continue
                    # Old format compat
                    prev_jd = (store_data.get("_gap_session") or {}).get("jd_text", "").strip()
                    if prev_jd == jd_text.strip():
                        skipped += 1
                        continue
                except Exception:
                    pass
            todo_ids.append(sid)

        # Create batch cancel flag (clean up old entries first)
        now_ms = int(time.time() * 1000)
        stale = [k for k in _batch_cancel_flags if int(k.split("_")[1]) < now_ms - 3600_000]
        for k in stale:
            _batch_cancel_flags.pop(k, None)
        batch_id = f"batch_{now_ms}"
        _batch_cancel_flags[batch_id] = False

        # Create jobs upfront (for frontend polling) but run sequentially
        created_jobs = []
        for sid in todo_ids:
            job = jobs.create(f"analyze_{sid[:8]}", anonymize=False, autofix=False, template_name="")
            created_jobs.append({"store_id": sid, "job_id": job.job_id})

        # Launch all batch analyze jobs — semaphore limits concurrency
        for item in created_jobs:
            sid, jid = item["store_id"], item["job_id"]
            cv_json = _load_store_cv(sid)
            if not cv_json:
                jobs.update(jid, status="Failed", progress=100, error="CV not found")
                continue
            thread = threading.Thread(
                target=_run_batch_analyze, args=(jid, sid, cv_json, jd_text, batch_id), daemon=True
            )
            thread.start()

        return {"ok": True, "jobs": created_jobs, "skipped": skipped, "batch_id": batch_id}

    if action not in ("generate", "anonymize", "tailor"):
        raise HTTPException(status_code=400, detail=f"Unknown action: {action}")

    if action == "tailor" and not jd_text.strip():
        raise HTTPException(status_code=400, detail="JD text is required for tailoring")

    if not template_name:
        tpls = sorted(TEMPLATES_DIR.glob("*.docx"))
        template_name = tpls[0].name if tpls else ""
    tpl_path = (TEMPLATES_DIR / template_name).resolve()
    if not str(tpl_path).startswith(str(TEMPLATES_DIR.resolve())) or not tpl_path.exists():
        raise HTTPException(status_code=400, detail=f"Unknown template: {template_name}")

    client_ip = request.client.host if request.client else "unknown"
    created_jobs = []

    for sid in ids:
        cv_json = _load_store_cv(sid)
        if not cv_json:
            continue
        workdir = make_temp_workspace()
        # Write a dummy source so the pipeline has a path
        dummy = workdir / "batch_cv.json"
        dummy.write_text(json.dumps(cv_json, ensure_ascii=False), encoding="utf-8")

        do_anon = anonymize or (action == "anonymize")
        do_tailor = action == "tailor"

        job = jobs.create(f"batch_{sid[:8]}.json", anonymize=do_anon, autofix=False, template_name=template_name)
        started_at = time.time()

        thread = threading.Thread(
            target=_run_job,
            args=(job.job_id, dummy, workdir, do_anon, False, do_tailor, jd_text, False, template_name, sid, client_ip, started_at, True),
            kwargs={"preloaded_data": cv_json},
            daemon=True,
        )
        thread.start()
        created_jobs.append({"store_id": sid, "job_id": job.job_id})

    return {"ok": True, "jobs": created_jobs}


def _resolve_github_token(cfg: dict) -> str:
    """Resolve GitHub PAT: env GITHUB_TOKEN > .github_token file > settings."""
    env = os.environ.get("GITHUB_TOKEN", "").strip()
    if env:
        return env
    token_file = APP_DIR / ".github_token"
    if token_file.exists():
        val = token_file.read_text(encoding="utf-8").strip()
        if val:
            return val
    return cfg.get("github_token", "")


@app.get("/setup", response_class=HTMLResponse)
def setup_page(request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    cfg = _core.load_config()
    current_key = resolve_api_key(APP_DIR, cfg)
    key_source = "not set"
    if os.environ.get("GEMINI_API_KEY", "").strip():
        key_source = "environment variable <code>GEMINI_API_KEY</code>"
    elif (APP_DIR / ".api_key").exists() and (APP_DIR / ".api_key").read_text().strip():
        key_source = "local <code>.api_key</code> file"
    elif cfg.get("gemini_api_key") or cfg.get("api_key"):
        key_source = "<code>~/.quantoricv_settings.json</code>"

    key_display = "[configured]" if current_key else "(not set)"
    status_color = "#2e7d32" if current_key else "#c62828"
    status_text = f"Key {key_display} — source: {key_source}" if current_key else "⚠️ No API key configured"

    # GitHub PAT status
    gh_token = _resolve_github_token(cfg)
    gh_display = "[configured]" if gh_token else "(not set)"
    gh_color = "#2e7d32" if gh_token else "#888"
    gh_source = ""
    if os.environ.get("GITHUB_TOKEN", "").strip():
        gh_source = "env <code>GITHUB_TOKEN</code>"
    elif (APP_DIR / ".github_token").exists():
        gh_source = "<code>.github_token</code> file"
    elif cfg.get("github_token"):
        gh_source = "<code>~/.quantoricv_settings.json</code>"
    gh_status = f"GitHub PAT {gh_display}" + (f" — source: {gh_source}" if gh_source else "")

    html = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Q-CV Setup</title>
  <style>
    body {{ font-family: Arial, sans-serif; max-width: 520px; margin: 60px auto; color: #222; }}
    h1 {{ margin-bottom: 4px; }}
    .status {{ padding: 10px 14px; border-radius: 6px; background: #f5f5f5;
               font-size: 14px; margin: 16px 0 24px; }}
    label {{ display: block; font-weight: bold; margin-bottom: 6px; }}
    input[type=text] {{ width: 100%; padding: 8px 10px; font-size: 14px; border: 1px solid #ccc;
                        border-radius: 4px; box-sizing: border-box; }}
    button {{ margin-top: 12px; padding: 9px 22px; background: #1565c0; color: #fff;
              border: none; border-radius: 4px; font-size: 14px; cursor: pointer; }}
    button:hover {{ background: #0d47a1; }}
    .note {{ margin-top: 20px; font-size: 12px; color: #666; }}
    a {{ color: #1565c0; }}
    .section {{ margin-bottom: 24px; }}
    hr {{ border: none; border-top: 1px solid #e0e0e0; margin: 24px 0; }}
  </style>
</head>
<body>
  <h1>Q-CV Setup</h1>
  <div class="status" style="color:{status_color}">{status_text}</div>
  <div class="section">
    <form method="post" action="/setup">
      <label for="key">Gemini API Key</label>
      <input type="text" id="key" name="api_key" placeholder="AIza..." autocomplete="off">
      <input type="hidden" name="field" value="gemini">
      <button type="submit">Save &amp; Apply</button>
    </form>
  </div>
  <hr>
  <div class="status" style="color:{gh_color}">{gh_status}</div>
  <div class="section">
    <form method="post" action="/setup">
      <label for="gh_key">GitHub Personal Access Token (for GitHub Miner)</label>
      <input type="text" id="gh_key" name="github_token" placeholder="ghp_..." autocomplete="off">
      <input type="hidden" name="field" value="github">
      <button type="submit">Save &amp; Apply</button>
    </form>
  </div>
  <p class="note">
    Keys are saved to local files and take effect immediately — no server restart needed.<br><br>
    Gemini priority: <code>GEMINI_API_KEY</code> env &gt; <code>.api_key</code> file &gt; <code>~/.quantoricv_settings.json</code><br>
    GitHub priority: <code>GITHUB_TOKEN</code> env &gt; <code>.github_token</code> file &gt; <code>~/.quantoricv_settings.json</code><br><br>
    <a href="/">← Back to converter</a>
  </p>
</body>
</html>"""
    return HTMLResponse(html)


@app.post("/setup")
async def setup_save(
    request: Request,
    api_key: str = Form(None),
    github_token: str = Form(None),
    field: str = Form("gemini"),
):
    _auth.require_role(_auth.ADMIN)(request)
    if field == "github" and github_token:
        token = github_token.strip()
        if not token:
            raise HTTPException(status_code=400, detail="Token cannot be empty")
        (APP_DIR / ".github_token").write_text(token, encoding="utf-8")
    elif api_key:
        key = api_key.strip()
        if not key:
            raise HTTPException(status_code=400, detail="API key cannot be empty")
        (APP_DIR / ".api_key").write_text(key, encoding="utf-8")
    return RedirectResponse(url="/setup?saved=1", status_code=303)


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    # If not authenticated, show login page
    user = _auth.get_current_user(request)
    if not user:
        login_path = APP_DIR / "templates" / "login.html"
        return RawResponse(
            content=login_path.read_bytes(),
            media_type="text/html",
            headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
        )
    index_path = APP_DIR / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=404, detail="index.html not found")
    return RawResponse(
        content=index_path.read_bytes(),
        media_type="text/html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


# ── Auth endpoints ──

@app.post("/api/auth/login")
async def auth_login(request: Request):
    body = await request.json()
    email = (body.get("email") or "").strip()
    password = body.get("password") or ""
    result = _auth.handle_login(email, password)
    if not result:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token, info = result
    response = RawResponse(
        content=json.dumps({"ok": True, **_auth.user_info_response(info)}),
        media_type="application/json",
    )
    response.set_cookie(
        _auth.COOKIE_NAME, token,
        httponly=True, samesite="lax", max_age=_auth.TOKEN_TTL,
    )
    return response


@app.get("/api/auth/me")
def auth_me(request: Request):
    user = _auth.get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return _auth.user_info_response(user)


@app.post("/api/auth/logout")
def auth_logout():
    response = RawResponse(
        content=json.dumps({"ok": True}),
        media_type="application/json",
    )
    response.delete_cookie(_auth.COOKIE_NAME)
    return response


@app.get("/api/auth/users")
def auth_list_users(request: Request):
    user = _auth.require_role(_auth.ADMIN)(request)
    return {"users": _auth.list_users()}


@app.put("/api/auth/users/{email}")
async def auth_upsert_user(email: str, request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    body = await request.json()
    _auth.upsert_user(
        email=email,
        name=body.get("name", ""),
        role=body.get("role", _auth.USER),
        password=body.get("password", ""),
        active=body.get("active", True),
    )
    return {"ok": True}


@app.delete("/api/auth/users/{email}")
def auth_delete_user(email: str, request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    _auth.delete_user(email)
    return {"ok": True}


@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(request: Request):
    _auth.require_role(_auth.ADMIN)(request)
    p = APP_DIR / "templates" / "admin_users.html"
    return RawResponse(content=p.read_bytes(), media_type="text/html")


@app.post("/admin/upload_data")
async def admin_upload_data(request: Request, file: UploadFile = File(...)):
    """Upload a zip archive to restore _store/, _jd_store/, _cache/ into DATA_DIR."""
    _auth.require_role(_auth.ADMIN)(request)
    import zipfile, io
    content = await file.read()
    if len(content) > 500 * 1024 * 1024:  # 500MB limit
        raise HTTPException(400, "File too large (max 500MB)")
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        raise HTTPException(400, "Invalid zip file")
    extracted = 0
    for name in zf.namelist():
        # Only allow known directories
        if not any(name.startswith(p) for p in ("_store/", "_jd_store/", "_cache/", "_users.json", "_employees.json", "_positions.json")):
            continue
        target = DATA_DIR / name
        if name.endswith("/"):
            target.mkdir(parents=True, exist_ok=True)
        else:
            target.parent.mkdir(parents=True, exist_ok=True)
            # Validate path doesn't escape DATA_DIR
            if not str(target.resolve()).startswith(str(DATA_DIR.resolve())):
                continue
            target.write_bytes(zf.read(name))
            extracted += 1
    # Reload caches
    global _embed_ids, _embed_matrix, _store_cache
    _load_embed_cache()
    _load_employees()
    _load_positions()
    # Force full cache rebuild
    global _store_cache_ready
    _store_cache = []
    _store_cache_ready = False
    _store_cache_init()
    return {"ok": True, "extracted": extracted}


@app.get("/admin/upload_data", response_class=HTMLResponse)
def admin_upload_page(request: Request):
    """Simple upload form for data restore."""
    _auth.require_role(_auth.ADMIN)(request)
    return RawResponse(content=b"""<!DOCTYPE html>
<html><head><title>Upload Data</title>
<style>body{font-family:sans-serif;padding:24px}h1{font-size:20px}.msg{margin-top:12px;font-size:14px}</style>
</head><body>
<h1>Upload Data Archive</h1>
<p style="color:#64748b;font-size:13px">Upload a .zip containing _store/, _jd_store/, _cache/ folders.</p>
<form id="f"><input type="file" name="file" accept=".zip" required>
<button type="submit" style="padding:8px 16px;margin-left:8px">Upload</button></form>
<div class="msg" id="msg"></div>
<p style="margin-top:16px"><a href="/">Back to app</a></p>
<script>
document.getElementById("f").addEventListener("submit", async function(e) {
  e.preventDefault();
  var fd = new FormData(this);
  document.getElementById("msg").textContent = "Uploading...";
  try {
    var r = await fetch("/admin/upload_data", {method:"POST", body: fd});
    var d = await r.json();
    document.getElementById("msg").textContent = r.ok
      ? "Done! Extracted " + d.extracted + " files."
      : "Error: " + (d.detail || "unknown");
  } catch(ex) { document.getElementById("msg").textContent = "Error: " + ex; }
});
</script></body></html>""", media_type="text/html")


# ── Employee Directory ────────────────────────────────────────────────────────

EMPLOYEES_PATH = DATA_DIR / "_employees.json"
_employees_cache: list[dict] = []

def _load_employees():
    global _employees_cache
    if EMPLOYEES_PATH.exists():
        _employees_cache = json.loads(EMPLOYEES_PATH.read_text(encoding="utf-8"))
    else:
        _employees_cache = []
    return _employees_cache

def _save_employees(data: list[dict]):
    global _employees_cache
    _employees_cache = data
    EMPLOYEES_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

_load_employees()


@app.post("/admin/upload_employees")
async def admin_upload_employees(request: Request, file: UploadFile = File(...)):
    """Upload employees.xlsx, parse and save as _employees.json."""
    _auth.require_role(_auth.ADMIN)(request)
    import openpyxl, io
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(400, "Invalid xlsx file")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Empty spreadsheet")
    headers = [str(h or "").strip() for h in rows[0]]
    employees = []
    for row in rows[1:]:
        rec = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                key = headers[i].lower()
                if hasattr(val, "isoformat"):
                    val = val.isoformat()[:10]
                rec[key] = val
        if rec.get("employee_name"):
            employees.append(rec)
    _save_employees(employees)
    return {"ok": True, "count": len(employees)}


@app.get("/admin/employees")
def list_employees(request: Request, q: str = ""):
    """Return employee directory, optionally filtered."""
    _auth.require_auth(request)
    data = _employees_cache or _load_employees()
    if q:
        q_lower = q.lower()
        data = [e for e in data if q_lower in json.dumps(e, ensure_ascii=False, default=str).lower()]
    return {"employees": data, "total": len(_employees_cache)}


# ── Positions / Projects Directory ───────────────────────────────────────────

POSITIONS_PATH = DATA_DIR / "_positions.json"
_positions_cache: list[dict] = []

def _load_positions():
    global _positions_cache
    if POSITIONS_PATH.exists():
        _positions_cache = json.loads(POSITIONS_PATH.read_text(encoding="utf-8"))
    else:
        _positions_cache = []
    return _positions_cache

def _save_positions(data: list[dict]):
    global _positions_cache
    _positions_cache = data
    POSITIONS_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

_load_positions()


@app.post("/admin/upload_positions")
async def admin_upload_positions(request: Request, file: UploadFile = File(...)):
    """Upload Position_Projects.xlsx, parse and save as _positions.json."""
    _auth.require_role(_auth.ADMIN)(request)
    import openpyxl, io
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(400, "Invalid xlsx file")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Empty spreadsheet")
    headers = [str(h or "").strip().lower() for h in rows[0]]
    positions = []
    for row in rows[1:]:
        rec = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                if hasattr(val, "isoformat"):
                    val = val.isoformat()[:10]
                rec[headers[i]] = val
        if rec.get("employee_name") or rec.get("position_code"):
            positions.append(rec)
    _save_positions(positions)
    return {"ok": True, "count": len(positions)}


@app.get("/admin/positions")
def list_positions(request: Request, q: str = "", employee: str = ""):
    """Return positions, optionally filtered by search or employee name."""
    _auth.require_auth(request)
    data = _positions_cache or _load_positions()
    if employee:
        emp_lower = employee.lower()
        data = [p for p in data if (p.get("employee_name") or "").lower() == emp_lower]
    if q:
        q_lower = q.lower()
        data = [p for p in data if q_lower in json.dumps(p, ensure_ascii=False, default=str).lower()]
    return {"positions": data, "total": len(_positions_cache)}


# ── JD Store ──────────────────────────────────────────────────────────────────

_JD_LOCK = threading.Lock()
_jd_cache: list[dict] = []
_jd_cache_ready = False
_jd_candidate_counts: dict[str, int] = {}  # jd_id -> candidate count
_jd_counts_ready = False

def _jd_cache_init():
    global _jd_cache_ready
    with _JD_LOCK:
        if _jd_cache_ready:
            return
        for p in JD_STORE_DIR.glob("*.json"):
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                _jd_cache.append(data)
            except Exception:
                continue
        _jd_cache.sort(key=lambda j: j.get("number", 0))
        _jd_cache_ready = True

def _jd_counts_init():
    """Build JD candidate counts from store files (once at startup)."""
    global _jd_counts_ready
    counts: dict[str, int] = {}
    for p in STORE_DIR.glob("*.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            for jid in data.get("_gap_sessions", {}):
                counts[jid] = counts.get(jid, 0) + 1
            old = data.get("_gap_session")
            if old and old.get("jd_text") and "_gap_sessions" not in data:
                old_jid = hashlib.sha256(old["jd_text"].strip().encode()).hexdigest()
                counts[old_jid] = counts.get(old_jid, 0) + 1
        except Exception:
            continue
    _jd_candidate_counts.update(counts)
    _jd_counts_ready = True


def _jd_next_number() -> int:
    if not _jd_cache:
        return 1
    return max(j.get("number", 0) for j in _jd_cache) + 1

def _jd_auto_title(text: str) -> str:
    """Try to extract a job title from the first few lines of JD text."""
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    if not lines:
        return "Untitled JD"
    # First non-empty line is often the title
    candidate = lines[0]
    # Strip common prefixes
    for prefix in ("job title:", "position:", "role:", "title:", "job description:", "jd:"):
        if candidate.lower().startswith(prefix):
            candidate = candidate[len(prefix):].strip()
    # If too long, truncate
    if len(candidate) > 80:
        candidate = candidate[:77] + "..."
    return candidate or "Untitled JD"

def _jd_auto_company(text: str) -> str:
    """Try to extract company name from JD text."""
    text_lower = text.lower()
    for marker in ("company:", "employer:", "organization:", "client:"):
        idx = text_lower.find(marker)
        if idx >= 0:
            rest = text[idx + len(marker):].strip()
            line = rest.split("\n")[0].strip().rstrip(".")
            if line:
                return line[:60]
    return ""


@app.get("/jd_store")
def list_jd_store(request: Request):
    _auth.require_auth(request)
    if not _jd_cache_ready:
        _jd_cache_init()
    if not _jd_counts_ready:
        _jd_counts_init()
    with _JD_LOCK:
        items = []
        for j in sorted(_jd_cache, key=lambda x: x.get("number", 0)):
            items.append({
                "id": j["id"],
                "number": j.get("number", 0),
                "title": j.get("title", ""),
                "company": j.get("company", ""),
                "date": j.get("date", ""),
                "candidates": _jd_candidate_counts.get(j["id"], 0),
            })
        return {"items": items}


@app.get("/jd_store/{jd_id}")
def get_jd_item(jd_id: str, request: Request):
    _auth.require_auth(request)
    _validate_store_id(jd_id)
    p = JD_STORE_DIR / f"{jd_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="JD not found")
    return json.loads(p.read_text(encoding="utf-8"))


@app.post("/jd_store")
async def create_jd(request: Request):
    _auth.require_auth(request)
    body = await request.json()
    text = body.get("text", "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="JD text is required")
    title = body.get("title", "").strip() or _jd_auto_title(text)
    company = body.get("company", "").strip() or _jd_auto_company(text)

    jd_id = hashlib.sha256(text.encode()).hexdigest()

    with _JD_LOCK:
        if not _jd_cache_ready:
            _jd_cache_init()
        # Check if JD with same text already exists
        for j in _jd_cache:
            if j["id"] == jd_id:
                return {"ok": True, "id": jd_id, "exists": True, "item": j}

        number = _jd_next_number()
        jd_data = {
            "id": jd_id,
            "number": number,
            "title": title,
            "company": company,
            "text": text,
            "date": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "candidates": 0,
        }
        (JD_STORE_DIR / f"{jd_id}.json").write_text(
            json.dumps(jd_data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        _jd_cache.append(jd_data)

    return {"ok": True, "id": jd_id, "exists": False, "item": jd_data}


@app.put("/jd_store/{jd_id}")
async def update_jd(jd_id: str, request: Request):
    _auth.require_auth(request)
    _validate_store_id(jd_id)
    body = await request.json()

    with _JD_LOCK:
        p = JD_STORE_DIR / f"{jd_id}.json"
        if not p.exists():
            raise HTTPException(status_code=404, detail="JD not found")
        data = json.loads(p.read_text(encoding="utf-8"))

        if "title" in body:
            data["title"] = body["title"]
        if "company" in body:
            data["company"] = body["company"]
        if "text" in body:
            data["text"] = body["text"]

        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        # Update cache
        for i, j in enumerate(_jd_cache):
            if j["id"] == jd_id:
                _jd_cache[i] = data
                break

    return {"ok": True, "item": data}


@app.delete("/jd_store/{jd_id}")
def delete_jd(jd_id: str, request: Request):
    _auth.require_auth(request)
    _validate_store_id(jd_id)
    with _JD_LOCK:
        p = JD_STORE_DIR / f"{jd_id}.json"
        if not p.exists():
            raise HTTPException(status_code=404, detail="JD not found")
        p.unlink()
        _jd_cache[:] = [j for j in _jd_cache if j["id"] != jd_id]
    return {"ok": True}


_online_users: dict[str, float] = {}  # email -> last_seen timestamp
_ONLINE_TIMEOUT = 30  # seconds — user is "online" if seen within this window

@app.get("/stats")
def server_stats(request: Request):
    user = _auth.require_auth(request)
    # Track online status
    _online_users[user["email"]] = time.time()

    today = time.strftime("%Y-%m-%d")
    # Use cached stats (refreshed every 30s to avoid parsing full log on every poll)
    global _stats_cache, _stats_cache_ts
    now = time.time()
    if now - _stats_cache_ts > 30 or _stats_cache.get("_date") != today:
        events = _read_usage_events()
        _stats_cache = {
            "_date": today,
            "today_done": sum(1 for e in events if e.get("event") == "done" and e.get("ts", "").startswith(today)),
            "today_failed": sum(1 for e in events if e.get("event") == "failed" and e.get("ts", "").startswith(today)),
            "total_done": sum(1 for e in events if e.get("event") == "done"),
        }
        _stats_cache_ts = now
    today_done = _stats_cache["today_done"]
    today_failed = _stats_cache["today_failed"]
    total_done = _stats_cache["total_done"]
    uptime_sec = int(time.time() - _SERVER_START)
    h, rem = divmod(uptime_sec, 3600)
    m, s = divmod(rem, 60)
    uptime_str = f"{h}h {m}m" if h else f"{m}m {s}s"
    result = {
        "active_jobs": jobs.active_count(),
        "today_processed": today_done,
        "today_failed": today_failed,
        "total_processed": total_done,
        "uptime": uptime_str,
    }
    # For admins — show who's online
    if user["role"] == _auth.ADMIN:
        now = time.time()
        result["online_users"] = [
            email for email, ts in _online_users.items()
            if now - ts < _ONLINE_TIMEOUT
        ]
    return result


@app.get("/templates")
def list_templates(request: Request):
    _auth.require_auth(request)
    if not TEMPLATES_DIR.exists():
        return {"templates": []}
    names = sorted([p.name for p in TEMPLATES_DIR.glob("*.docx") if p.is_file()])
    return {"templates": names}


def build_source_key(source_path: Path) -> str:
    h = hashlib.sha256()
    with source_path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _build_processing_details(
    source_name: str,
    source_path: Path,
    template_name: str,
    anonymize: bool,
    autofix: bool,
    output_path: Path | None = None,
    content_details: dict | None = None,
) -> dict:
    suffix = source_path.suffix.lower()
    source_type = {
        ".pdf": "PDF",
        ".docx": "DOCX",
        ".png": "PNG",
        ".jpg": "JPG",
        ".jpeg": "JPEG",
    }.get(suffix, suffix.lstrip(".").upper() or "Unknown")

    details = {
        "source_type": source_type,
        "source_file": source_name,
        "template": template_name,
        "anonymize": bool(anonymize),
        "autofix": bool(autofix),
        "reuse_enabled": True,
        "image_input": suffix in {".png", ".jpg", ".jpeg"},
        "output_generated": output_path is not None,
        "output_file": output_path.name if output_path else None,
    }
    if content_details:
        details["content_details"] = content_details
    return details


## ── CV Store helpers ─────────────────────────────────────────────────

# In-memory cache of _meta dicts — avoids reading all JSON files on every /store request
_store_cache: list[dict] = []
_store_cache_ready = False

def _store_cache_init():
    """Load all _meta from store files into cache. Called once at startup."""
    global _store_cache_ready
    with _STORE_LOCK:
        existing_ids = {m.get("id") for m in _store_cache}
        for p in STORE_DIR.glob("*.json"):
            if p.stem in existing_ids:
                continue
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                meta = data.get("_meta", {})
                meta["id"] = p.stem
                _store_cache.append(meta)
            except Exception:
                continue
        _store_cache.sort(key=lambda m: m.get("date", ""), reverse=True)
        _store_cache_ready = True

def _store_cache_refresh():
    """Pick up new/changed store files (e.g. by employee_scanner or gap analysis)."""
    with _STORE_LOCK:
        cached_map = {m.get("id"): m for m in (_store_cache or [])}
        changed = 0
        for p in STORE_DIR.glob("*.json"):
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                meta = data.get("_meta", {})
                meta["id"] = p.stem
                old = cached_map.get(p.stem)
                if not old:
                    _store_cache.append(meta)
                    changed += 1
                elif old.get("date") != meta.get("date") or old.get("match_pct") != meta.get("match_pct"):
                    # Update existing entry
                    for i, m in enumerate(_store_cache):
                        if m.get("id") == p.stem:
                            _store_cache[i] = meta
                            break
                    changed += 1
            except Exception:
                continue
        if changed:
            _store_cache.sort(key=lambda m: m.get("date", ""), reverse=True)
        return changed

def _store_cache_upsert(meta: dict):
    """Add or update a meta entry in the cache. Must be called with _STORE_LOCK held."""
    sid = meta.get("id", "")
    for i, m in enumerate(_store_cache):
        if m.get("id") == sid:
            _store_cache[i] = meta
            return
    _store_cache.append(meta)

def _store_cache_remove(store_id: str):
    """Remove an entry from the cache. Thread-safe."""
    with _STORE_LOCK:
        global _store_cache
        _store_cache = [m for m in _store_cache if m.get("id") != store_id]

def _store_cache_get_meta(store_id: str) -> dict | None:
    """Get cached meta by ID. Thread-safe."""
    with _STORE_LOCK:
        for m in _store_cache:
            if m.get("id") == store_id:
                return dict(m)
    return None


def _find_store_by_name(name: str) -> Path | None:
    """Check if a CV with this name already exists in store (uses cache)."""
    if not name:
        return None
    name_lower = name.lower()
    # Try cache first (fast path)
    if _store_cache_ready:
        with _STORE_LOCK:
            for m in _store_cache:
                if m.get("name", "").lower() == name_lower:
                    p = STORE_DIR / f"{m['id']}.json"
                    if p.exists():
                        return p
        return None
    # Fallback: read files (only during startup before cache is ready)
    for p in STORE_DIR.glob("*.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            if data.get("_meta", {}).get("name", "").lower() == name_lower:
                return p
        except Exception:
            continue
    return None


def _save_to_store(store_id: str, cv_json: dict, source_filename: str) -> None:
    """Persist extracted CV JSON with metadata to _store/."""
    with _STORE_LOCK:
        basics = cv_json.get("basics", {})
        existing = _find_store_by_name(basics.get("name", ""))
        if existing and existing.stem != store_id:
            return
        exp = cv_json.get("experience", [])
        # Auto-detect source (like desktop Q-CV)
        fname_lower = (source_filename or "").lower()
        links_dump = json.dumps(basics.get("links", [])).lower()
        if "linkedin.com" in links_dump or "linkedin" in fname_lower or fname_lower.startswith("profile"):
            comments = "Source: LinkedIn"
        elif "github.com" in fname_lower:
            comments = "Source: GitHub"
        else:
            comments = ""

        # Build search index (like desktop Q-CV: name, title, company, skills, filename, comments)
        skills_text = json.dumps(cv_json.get("skills", {}), ensure_ascii=False).lower()
        search_text = " ".join([
            basics.get("name", ""),
            basics.get("current_title", ""),
            exp[0].get("company_name", "") if exp else "",
            source_filename or "",
            comments,
            skills_text,
        ]).lower()

        meta = {
            "id": store_id,
            "name": basics.get("name", ""),
            "role": basics.get("current_title", ""),
            "company": exp[0].get("company_name", "") if exp else "",
            "date": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "source_filename": source_filename,
            "comments": comments,
            "search_text": search_text,
        }
        data = {"_meta": meta, **{k: v for k, v in cv_json.items() if k != "_meta"}}
        (STORE_DIR / f"{store_id}.json").write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        _store_cache_upsert(dict(meta))
        # Compute embedding in background (non-blocking)
        if store_id not in _embed_ids:
            threading.Thread(target=_embed_cv_async, args=(store_id, cv_json), daemon=True).start()


def _embed_cv_async(store_id: str, cv_json: dict):
    """Compute and store embedding for a CV (runs in background thread)."""
    text = _cv_text_for_embedding(cv_json)
    if text.strip():
        vec = _compute_embedding(text)
        if vec:
            _add_embedding(store_id, vec)


def _jd_id_for_text(jd_text: str) -> str:
    """Compute JD ID from text (sha256). Also ensures JD exists in JD Store."""
    jd_id = hashlib.sha256(jd_text.strip().encode()).hexdigest()
    # Auto-create JD in JD Store if not exists
    jd_path = JD_STORE_DIR / f"{jd_id}.json"
    if not jd_path.exists():
        with _JD_LOCK:
            if not _jd_cache_ready:
                _jd_cache_init()
            if not any(j["id"] == jd_id for j in _jd_cache):
                jd_data = {
                    "id": jd_id,
                    "number": _jd_next_number(),
                    "title": _jd_auto_title(jd_text),
                    "company": _jd_auto_company(jd_text),
                    "text": jd_text.strip(),
                    "date": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "candidates": 0,
                }
                jd_path.write_text(json.dumps(jd_data, ensure_ascii=False, indent=2), encoding="utf-8")
                _jd_cache.append(jd_data)
    return jd_id


def _save_store_gap(store_id: str, gap_analysis: dict, jd_text: str, base_json: dict = None) -> None:
    """Save gap analysis to store entry — supports multiple JDs via _gap_sessions."""
    jd_id = _jd_id_for_text(jd_text)
    with _STORE_LOCK:
        p = STORE_DIR / f"{store_id}.json"
        if not p.exists():
            name = (base_json or {}).get("basics", {}).get("name", "")
            p = _find_store_by_name(name)
            if not p:
                return
        data = json.loads(p.read_text(encoding="utf-8"))

        # Migrate old format if needed
        if "_gap_session" in data and "_gap_sessions" not in data:
            old = data.pop("_gap_session")
            old_jd_id = hashlib.sha256(old.get("jd_text", "").strip().encode()).hexdigest()
            data["_gap_sessions"] = {old_jd_id: old}

        if "_gap_sessions" not in data:
            data["_gap_sessions"] = {}

        data["_gap_sessions"][jd_id] = {
            "gap_analysis": gap_analysis,
            "jd_text": jd_text,
            "date": time.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        # Also keep _gap_session for backward compat with frontend (points to latest)
        data["_gap_session"] = data["_gap_sessions"][jd_id]

        data["_meta"]["analyzed"] = True
        data["_meta"]["tailored"] = False
        data.pop("_tailor_session", None)
        data["_meta"]["match_pct"] = int(gap_analysis.get("match_percentage", 0))
        data["_meta"]["match_jd_id"] = jd_id
        data["_meta"]["date"] = time.strftime("%Y-%m-%dT%H:%M:%S")
        data["_meta"]["id"] = p.stem
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        _store_cache_upsert(dict(data["_meta"]))
        # Update JD candidate count
        _jd_candidate_counts[jd_id] = _jd_candidate_counts.get(jd_id, 0) + 1


def _update_store_tailor(store_id: str, tailored_json: dict, jd_text: str,
                         gap_analysis: dict, focus_skills: list,
                         keyword_report: dict) -> None:
    """Update an existing store entry with tailoring session data."""
    with _STORE_LOCK:
        p = STORE_DIR / f"{store_id}.json"
        if not p.exists():
            name = tailored_json.get("basics", {}).get("name", "")
            p = _find_store_by_name(name)
            if not p:
                return
        data = json.loads(p.read_text(encoding="utf-8"))
        data["_tailor_session"] = {
            "tailored_json": tailored_json,
            "jd_text": jd_text,
            "gap_analysis": gap_analysis,
            "focus_skills": focus_skills,
            "keyword_report": keyword_report,
            "date": time.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        data["_meta"]["tailored"] = True
        data["_meta"]["date"] = time.strftime("%Y-%m-%dT%H:%M:%S")
        # Keep gap analysis match_percentage (consistent with Fit Report on Convert tab)
        if gap_analysis and "match_percentage" in gap_analysis:
            data["_meta"]["match_pct"] = int(gap_analysis["match_percentage"])
        data["_meta"]["id"] = p.stem
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        _store_cache_upsert(dict(data["_meta"]))


_store_cache_last_refresh = 0.0

def _list_store() -> list[dict]:
    """Return list of _meta dicts for all stored CVs (from cache)."""
    global _store_cache_last_refresh
    if not _store_cache_ready:
        _store_cache_init()
    # Pick up externally-added files every 10 seconds
    now = time.time()
    if now - _store_cache_last_refresh > 10:
        _store_cache_last_refresh = now
        _store_cache_refresh()
    with _STORE_LOCK:
        return sorted(list(_store_cache or []), key=lambda m: m.get("date", ""), reverse=True)


def _load_store_cv(store_id: str) -> dict | None:
    """Load a stored CV JSON, stripping _meta."""
    p = STORE_DIR / f"{store_id}.json"
    if not p.exists():
        return None
    data = json.loads(p.read_text(encoding="utf-8"))
    data.pop("_meta", None)
    return data


def _run_batch_analyze(job_id: str, store_id: str, cv_json: dict, jd_text: str, batch_id: str = "") -> None:
    """Lightweight batch analysis: gap analysis only, no DOCX generation."""
    def _is_cancelled():
        if batch_id and _batch_cancel_flags.get(batch_id):
            return True
        job = jobs.get(job_id)
        return job and getattr(job, "_cancelled", False)

    # Wait for semaphore, checking cancel every second
    while not _JOB_SEMAPHORE.acquire(timeout=1):
        if _is_cancelled():
            jobs.update(job_id, status="Cancelled", progress=100)
            return
    try:
        if _is_cancelled():
            jobs.update(job_id, status="Cancelled", progress=100)
            return

        jobs.update(job_id, status="Analyzing", progress=30)
        engine = QCVWebEngine(TEMPLATES_DIR)
        engine.config = _core.load_config()
        engine.model_name = choose_model_name(engine.config)
        api_key = resolve_api_key(engine.app_dir, engine.config)
        configure_gemini(api_key)

        gap_result = engine._analyze_gap(cv_json, jd_text)
        gap_result["_output_base"] = _build_output_base_name(cv_json, anonymize=False)

        _save_store_gap(store_id, gap_result, jd_text, cv_json)

        job = jobs.get(job_id)
        if job:
            setattr(job, "_gap_analysis", gap_result)
        jobs.update(job_id, status="Done", progress=100)
    except Exception as e:
        jobs.update(job_id, status="Failed", progress=100, error=str(e))
    finally:
        _JOB_SEMAPHORE.release()


def _build_projects_section(cv_data: dict) -> list[dict] | None:
    """Build 'Projects (Quantori Staffing)' section from positions data."""
    if not _positions_cache:
        return None
    name = (cv_data.get("basics") or {}).get("name", "")
    if not name:
        return None
    name_lower = name.lower()
    positions = [p for p in _positions_cache if (p.get("employee_name") or "").lower() == name_lower]
    if not positions:
        return None
    # Sort: active first (Assigned/Booked), then by date desc
    status_order = {"Assigned": 0, "Booked": 1, "Proposed": 2, "Requested": 3, "Ended": 4}
    # Sort: active statuses first, then by close date descending (most recent first)
    positions.sort(key=lambda p: (
        status_order.get(p.get("workload_status", ""), 9),
        "".join(chr(255 - ord(c)) for c in (p.get("wd_close_date") or "0000-00-00")),
    ))
    items = []
    for p in positions:
        account = p.get("account", "")
        role = p.get("position_code", "")
        status = p.get("workload_status", "")
        start = (p.get("wd_open_date") or "")[:7]  # YYYY-MM
        end = (p.get("wd_close_date") or "")[:7]
        fte = p.get("fte", "")
        period = f"{start}" + (f" - {end}" if end and end != start else "")
        line = f"{account} — {role}"
        if fte and fte != 1:
            line += f" ({fte} FTE)"
        line += f", {status}, {period}"
        items.append(line)
    return [{"title": "Projects (Quantori Staffing)", "items": items}]


def _run_job(job_id: str, source_path: Path, workdir: Path, anonymize: bool, autofix: bool, tailor: bool, jd_text: str, force_tailor: bool, template_name: str, source_key: str | None, client_ip: str, started_at: float, skip_gap: bool = False, preloaded_focus_skills: list | None = None, preloaded_data: dict | None = None, preloaded_gap: dict | None = None, user_email: str = "anonymous", add_projects: bool = False) -> None:
    jobs.update(job_id, status="Queued", progress=0)
    _JOB_SEMAPHORE.acquire()
    try:
        def cb(status: str, progress: int) -> None:
            jobs.update(job_id, status=status, progress=progress)

        def dbg(text: str) -> None:
            jobs.update(job_id, debug=text)

        # Shared store ID — set by early save, reused by final save
        _resolved_store_id = [None]  # mutable container for closure

        # Create pause event for gap analysis (tailor jobs only, unless skip_gap)
        pause_event = None
        gap_ready_cb = None
        focus_skills_cb = None
        if tailor and jd_text.strip() and not skip_gap:
            _raw_pause = threading.Event()
            class _SemaphorePause:
                """Wrapper that releases semaphore while waiting for user."""
                def wait(self, timeout=None):
                    _JOB_SEMAPHORE.release()
                    try:
                        return _raw_pause.wait(timeout=timeout)
                    finally:
                        _JOB_SEMAPHORE.acquire()
                def set(self):
                    _raw_pause.set()
                def is_set(self):
                    return _raw_pause.is_set()
            pause_event = _SemaphorePause()

            def gap_ready_cb(gap_result: dict, base_json: dict = None) -> None:
                job = jobs.get(job_id)
                if job:
                    setattr(job, "_gap_analysis", gap_result)
                    setattr(job, "_pause_event", pause_event)
                    if base_json:
                        setattr(job, "_cv_json", base_json)
                        # Early save to store — CV available before DOCX generation
                        try:
                            sid = source_key or hashlib.sha256(
                                json.dumps(base_json, sort_keys=True).encode()
                            ).hexdigest()
                            # Dedup by name
                            name = base_json.get("basics", {}).get("name", "")
                            existing = _find_store_by_name(name)
                            if existing:
                                sid = existing.stem
                            _resolved_store_id[0] = sid
                            if not (STORE_DIR / f"{sid}.json").exists():
                                _save_to_store(sid, base_json, source_path.name)
                            # Save gap analysis immediately (shows "Analyzed" badge)
                            _save_store_gap(sid, gap_result, jd_text, base_json)
                        except Exception:
                            pass

            def focus_skills_cb() -> list:
                job = jobs.get(job_id)
                if job and getattr(job, "_cancelled", False):
                    raise RuntimeError("Job cancelled by user")
                return getattr(job, "_focus_skills", []) if job else []

        # skip_gap: pre-fill focus_skills and gap_analysis, no pause
        if skip_gap and tailor:
            _preloaded_fs = preloaded_focus_skills or []
            def focus_skills_cb() -> list:
                return _preloaded_fs
            # Preserve gap_analysis and focus_skills on job for store save
            job = jobs.get(job_id)
            if job:
                if preloaded_gap:
                    setattr(job, "_gap_analysis", preloaded_gap)
                setattr(job, "_focus_skills", _preloaded_fs)

        job_engine = QCVWebEngine(TEMPLATES_DIR)
        result_path = job_engine.process(
            source_path=source_path,
            output_dir=workdir,
            anonymize=anonymize,
            autofix=autofix,
            tailor=tailor,
            jd_text=jd_text,
            force_tailor=force_tailor,
            template_name=template_name,
            source_key=source_key,
            status_cb=cb,
            debug_cb=dbg,
            pause_event=pause_event,
            gap_ready_cb=gap_ready_cb,
            focus_skills_cb=focus_skills_cb,
            preloaded_data=preloaded_data,
            extra_sections_cb=_build_projects_section if add_projects else None,
        )

        # Store base CV JSON on job for download
        job = jobs.get(job_id)
        base_json = getattr(job_engine, "_last_base_json", None)
        if base_json and job:
            setattr(job, "_cv_json", base_json)

        if job:
            details = _build_processing_details(
                source_name=getattr(job, "filename", source_path.name),
                source_path=source_path,
                template_name=template_name,
                anonymize=anonymize,
                autofix=autofix,
                output_path=result_path,
                content_details=getattr(job_engine, "last_content_details", None),
            )
            setattr(job, "details", details)
            # Store data for potential refine pass
            if tailor and jd_text.strip():
                setattr(job, "_tailored_json", getattr(job_engine, "_last_tailored_json", None))
                setattr(job, "_jd_text", jd_text)
                setattr(job, "_output_dir", str(workdir))
                setattr(job, "_source_name", source_path.name)

        # Check if job was cancelled while running
        job = jobs.get(job_id)
        if job and getattr(job, "_cancelled", False):
            jobs.update(job_id, status="Cancelled", progress=100)
            return
        jobs.update(job_id, status="Done", progress=100, result_path=str(result_path))

        # Auto-save base CV JSON to persistent store
        try:
            job = jobs.get(job_id)
            if base_json:
                # Reuse store ID from early save if available
                sid = _resolved_store_id[0]
                if not sid:
                    sid = source_key or hashlib.sha256(
                        json.dumps(base_json, sort_keys=True).encode()
                    ).hexdigest()
                    name = base_json.get("basics", {}).get("name", "")
                    existing_by_name = _find_store_by_name(name)
                    if existing_by_name:
                        sid = existing_by_name.stem
                if not (STORE_DIR / f"{sid}.json").exists():
                    _save_to_store(sid, base_json, source_path.name)
                # Save tailoring session if tailor was performed
                if tailor and jd_text.strip():
                    tailored = getattr(job_engine, "_last_tailored_json", None)
                    gap = getattr(job, "_gap_analysis", None) if job else None
                    focus = getattr(job, "_focus_skills", []) if job else []
                    cd = (getattr(job, "details", None) or {}).get("content_details") or {}
                    kw_report = cd.get("jd_keyword_report", {})
                    if tailored:
                        _update_store_tailor(sid, tailored, jd_text, gap or {}, focus, kw_report)
        except Exception:
            pass

        append_usage({
            "event": "done",
            "job_id": job_id,
            "ip": client_ip,
            "user": user_email,
            "file": source_path.name,
            "template": template_name,
            "anonymize": anonymize,
            "autofix": autofix,
            "tailor": tailor,
            "duration_sec": round(time.time() - started_at, 2),
        })
    except LowRelevanceError as e:
        jobs.update(job_id, status="Low Relevance", progress=100, error=str(e))
        append_usage({
            "event": "skipped_low_relevance",
            "job_id": job_id, "ip": client_ip, "user": user_email,
            "file": source_path.name, "tailor": True,
            "duration_sec": round(time.time() - started_at, 2),
        })
    except Exception as e:
        jobs.update(job_id, status="Failed", progress=100, error=str(e))
        append_usage({
            "event": "failed",
            "job_id": job_id,
            "ip": client_ip,
            "user": user_email,
            "file": source_path.name,
            "template": template_name,
            "anonymize": anonymize,
            "autofix": autofix,
            "tailor": tailor,
            "duration_sec": round(time.time() - started_at, 2),
            "error": str(e),
        })
    finally:
        _JOB_SEMAPHORE.release()


@app.post("/jobs")
async def create_job(
    request: Request,
    file: UploadFile = File(...),
    anonymize: bool = Form(False),
    autofix: bool = Form(False),
    tailor: bool = Form(False),
    jd_text: str = Form(""),
    template_name: str = Form(...),
    force_tailor: bool = Form(False),
    skip_gap: bool = Form(False),
    focus_skills_json: str = Form(""),
    import_only: bool = Form(False),
    store_id: str = Form(""),
    add_projects: bool = Form(False),
):
    _auth.require_auth(request)
    suffix = Path(file.filename or "upload.docx").suffix.lower()
    if suffix not in {".pdf", ".docx", ".png", ".jpg", ".jpeg", ".json"}:
        raise HTTPException(status_code=400, detail="Only PDF, DOCX, PNG, JPG, JPEG, and JSON are supported.")

    if not template_name:
        raise HTTPException(status_code=400, detail="Template is required.")

    template_path = (TEMPLATES_DIR / template_name).resolve()
    if not str(template_path).startswith(str(TEMPLATES_DIR.resolve())) or not template_path.exists():
        raise HTTPException(status_code=400, detail=f"Unknown template: {template_name}")

    # Read uploaded file
    workdir = make_temp_workspace()
    source_path = workdir / (file.filename or "uploaded_file")
    with source_path.open("wb") as f:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            f.write(chunk)

    # Handle JSON CV upload: extract CV data and optional _fit_session
    preloaded_data = None
    fit_session = None
    if suffix == ".json":
        try:
            raw = json.loads(source_path.read_text(encoding="utf-8"))
            fit_session = raw.pop("_fit_session", None)
            preloaded_data = raw  # remaining dict is the CV JSON
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise HTTPException(status_code=400, detail=f"Invalid JSON file: {exc}")

        # If _fit_session contains JD and user didn't provide one, use it
        if fit_session and not jd_text.strip():
            jd_text = fit_session.get("jd_text", "")
        # Auto-enable tailor if fit_session has JD — but only if user didn't explicitly disable it
        if fit_session and fit_session.get("jd_text", "").strip() and tailor:
            pass  # keep tailor=True (user has checkbox on)
        elif not tailor:
            fit_session = None  # user unchecked tailor — ignore fit_session
        # Extract focus_skills from fit_session (only when frontend requested skip_gap)
        if fit_session and skip_gap:
            user_edits = fit_session.get("user_edits", {})
            if not focus_skills_json and user_edits.get("checked_skills"):
                focus_skills_json = json.dumps(user_edits["checked_skills"])

    if tailor and not jd_text.strip():
        raise HTTPException(status_code=400, detail="Job description is required when tailoring is enabled.")

    source_key = build_source_key(source_path) if suffix != ".json" else None
    # For CVs loaded from store (JSON), use store_id as source_key
    if not source_key and store_id:
        source_key = store_id

    # Skip if already in store (batch import dedup only)
    if import_only and source_key and (STORE_DIR / f"{source_key}.json").exists():
        return {"job_id": "skip", "status": "Done", "progress": 100, "already_in_store": True}

    job = jobs.create(
        file.filename or "uploaded_file",
        anonymize=anonymize,
        autofix=autofix,
        template_name=template_name,
    )

    details = _build_processing_details(
        source_name=file.filename or source_path.name,
        source_path=source_path,
        template_name=template_name,
        anonymize=anonymize,
        autofix=autofix,
        output_path=None,
        content_details=None,
    )
    setattr(job, "details", details)

    client_ip = request.client.host if request.client else "unknown"
    user = _auth.get_current_user(request)
    user_email = user["email"] if user else "anonymous"
    setattr(job, "_user_email", user_email)
    started_at = time.time()
    append_usage({
        "event": "started",
        "job_id": job.job_id,
        "ip": client_ip,
        "user": user_email,
        "file": source_path.name,
        "template": template_name,
        "anonymize": anonymize,
        "autofix": autofix,
        "tailor": tailor,
        "size_bytes": source_path.stat().st_size if source_path.exists() else None,
    })

    thread = threading.Thread(
        target=_run_job,
        args=(job.job_id, source_path, workdir, anonymize, autofix, tailor, jd_text, force_tailor, template_name, source_key, client_ip, started_at, skip_gap),
        kwargs={
            "preloaded_focus_skills": json.loads(focus_skills_json) if focus_skills_json else None,
            "preloaded_data": preloaded_data,
            "preloaded_gap": fit_session.get("gap_analysis") if fit_session else None,
            "user_email": user_email,
            "add_projects": add_projects,
        },
        daemon=True,
    )
    thread.start()

    return {
        "job_id": job.job_id,
        "status": job.status,
        "progress": job.progress,
        "filename": job.filename,
        "template": template_name,
        "details": getattr(job, "details", None),
    }


@app.get("/jobs/{job_id}")
def get_job(job_id: str, request: Request):
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    resp = {
        "job_id": job.job_id,
        "filename": job.filename,
        "status": job.status,
        "progress": job.progress,
        "error": job.error,
        "debug": getattr(job, "debug", ""),
        "ready": bool(job.result_path),
        "details": getattr(job, "details", None),
    }
    gap = getattr(job, "_gap_analysis", None)
    if gap:
        resp["gap_analysis"] = gap
    return resp


@app.get("/jobs/{job_id}/cv_json")
def get_cv_json(job_id: str, request: Request):
    """Return the extracted base CV JSON for this job."""
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    cv_json = getattr(job, "_cv_json", None)
    if not cv_json:
        raise HTTPException(status_code=404, detail="CV JSON not available yet")
    return cv_json


@app.put("/jobs/{job_id}/cv_json")
async def update_cv_json(job_id: str, request: Request):
    """Update the base CV JSON for this job (from the editor)."""
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    body = await request.json()
    # Backup previous version
    prev = getattr(job, "_cv_json", None)
    if prev:
        bak_list = getattr(job, "_cv_json_bak", [])
        bak_list.append(copy.deepcopy(prev))
        setattr(job, "_cv_json_bak", bak_list)
    setattr(job, "_cv_json", body)
    return {"ok": True}


@app.post("/jobs/{job_id}/reanalyze")
async def reanalyze_job(job_id: str, request: Request):
    """Re-run gap analysis on the (possibly edited) CV JSON."""
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    cv_json = getattr(job, "_cv_json", None)
    if not cv_json:
        raise HTTPException(status_code=400, detail="No CV JSON available")
    body = await request.json()
    jd_text = body.get("jd_text", "")
    if not jd_text.strip():
        raise HTTPException(status_code=400, detail="Job description is required")

    import asyncio

    def _do_reanalyze():
        _JOB_SEMAPHORE.acquire()
        try:
            engine = QCVWebEngine(TEMPLATES_DIR)
            engine.config = _core.load_config()
            engine.model_name = choose_model_name(engine.config)
            api_key = resolve_api_key(engine.app_dir, engine.config)
            configure_gemini(api_key)
            gap_result = engine._analyze_gap(cv_json, jd_text)
            gap_result["_output_base"] = _build_output_base_name(cv_json, anonymize=False)
            setattr(job, "_gap_analysis", gap_result)
            return gap_result
        finally:
            _JOB_SEMAPHORE.release()

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _do_reanalyze)


@app.get("/jobs/{job_id}/download")
def download_job_result(job_id: str, request: Request):
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status != "Done" or not job.result_path:
        raise HTTPException(status_code=400, detail="Result is not ready yet")

    result_path = Path(job.result_path)
    if not result_path.exists():
        raise HTTPException(status_code=404, detail="Output file missing")

    return FileResponse(
        path=result_path,
        filename=result_path.name,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


def _run_refine(job_id: str, tailored_json: dict, jd_text: str, missing_keywords: list[str],
                output_dir: str, anonymize: bool, template_name: str, source_name: str,
                client_ip: str, started_at: float) -> None:
    _JOB_SEMAPHORE.acquire()
    try:
        def cb(status: str, progress: int) -> None:
            jobs.update(job_id, status=status, progress=progress)

        def dbg(text: str) -> None:
            jobs.update(job_id, debug=text)

        engine = QCVWebEngine(TEMPLATES_DIR)
        result_path = engine.refine(
            tailored_json=tailored_json,
            jd_text=jd_text,
            missing_keywords=missing_keywords,
            output_dir=Path(output_dir),
            anonymize=anonymize,
            template_name=template_name,
            source_name=source_name,
            status_cb=cb,
            debug_cb=dbg,
        )

        job = jobs.get(job_id)
        if job:
            details = _build_processing_details(
                source_name=source_name,
                source_path=Path(source_name),
                template_name=template_name,
                anonymize=anonymize,
                autofix=False,
                output_path=result_path,
                content_details=getattr(engine, "last_content_details", None),
            )
            setattr(job, "details", details)
            # Update stored tailored JSON for potential further refines
            setattr(job, "_tailored_json", getattr(engine, "_last_tailored_json", None) or tailored_json)

        jobs.update(job_id, status="Done", progress=100, result_path=str(result_path))
    except Exception as e:
        jobs.update(job_id, status="Failed", progress=100, error=str(e))
    finally:
        _JOB_SEMAPHORE.release()


@app.post("/jobs/{job_id}/refine")
async def refine_job(job_id: str, request: Request):
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status != "Done":
        raise HTTPException(status_code=400, detail="Job must be in Done state to refine")

    tailored_json = getattr(job, "_tailored_json", None)
    jd_text = getattr(job, "_jd_text", None)
    output_dir = getattr(job, "_output_dir", None)
    source_name = getattr(job, "_source_name", "refined")

    if not tailored_json or not jd_text:
        raise HTTPException(status_code=400, detail="No tailoring data available for this job")

    # Get missing keywords from current keyword report
    details = getattr(job, "details", None) or {}
    cd = details.get("content_details") or {}
    kw_report = cd.get("jd_keyword_report") or {}
    missing = kw_report.get("missing", [])
    if not missing:
        raise HTTPException(status_code=400, detail="No missing keywords to refine")

    # Reset job status for refine pass — clear result_path so polling doesn't see it as ready
    jobs.update(job_id, status="Refining", progress=0, error=None, result_path="")

    client_ip = request.client.host if request.client else "unknown"
    thread = threading.Thread(
        target=_run_refine,
        args=(job_id, tailored_json, jd_text, missing, output_dir,
              job.anonymize, job.template_name, source_name,
              client_ip, time.time()),
        daemon=True,
    )
    thread.start()

    return {"job_id": job_id, "status": "Refining"}


@app.post("/jobs/{job_id}/continue")
async def continue_job(job_id: str, request: Request):
    """Unblock a job paused at gap_analysis_ready to proceed with tailoring."""
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status != "gap_analysis_ready":
        raise HTTPException(status_code=400, detail="Job is not waiting for continuation")

    pause_event = getattr(job, "_pause_event", None)
    if not pause_event:
        raise HTTPException(status_code=400, detail="No pause event found")

    # Store focus skills selected by user for the tailor prompt
    try:
        body = await request.json()
        focus_skills = body.get("focus_skills", [])
    except Exception:
        focus_skills = []
    if focus_skills:
        setattr(job, "_focus_skills", focus_skills)
    pause_event.set()
    return {"job_id": job_id, "status": "Resuming"}


@app.post("/jobs/{job_id}/cancel")
async def cancel_job(job_id: str, request: Request):
    """Cancel a pending job — unblocks pause_event so the thread can exit."""
    _auth.require_auth(request)
    job = jobs.get(job_id)
    if not job:
        return {"ok": True}  # already gone
    # Set cancelled flag so gap analysis raises after unblock
    setattr(job, "_cancelled", True)
    # Unblock pause_event if waiting
    pause_event = getattr(job, "_pause_event", None)
    if pause_event:
        pause_event.set()
    # Mark as failed
    if job.status not in ("Done", "Failed"):
        jobs.update(job_id, status="Failed", progress=100, error="Cancelled by user")
    return {"ok": True}


@app.post("/batch/{batch_id}/cancel")
async def cancel_batch(batch_id: str, request: Request):
    """Cancel all pending jobs in a batch."""
    _auth.require_auth(request)
    if batch_id in _batch_cancel_flags:
        _batch_cancel_flags[batch_id] = True
    return {"ok": True}


@app.post("/xray")
async def xray_search(request: Request):
    """Generate X-Ray Boolean search queries from a candidate description."""
    _auth.require_auth(request)
    body = await request.json()
    user_input = (body.get("query") or "").strip()
    if len(user_input) < 5:
        raise HTTPException(400, "Please enter a longer description (at least 5 characters).")
    config = _core.load_config()
    model_name = choose_model_name(config)
    api_key = resolve_api_key(APP_DIR, config)
    configure_gemini(api_key)
    prompt_template = config.get("prompt_xray", _core.DEFAULT_PROMPTS.get("prompt_xray", ""))
    if not prompt_template:
        raise HTTPException(500, "X-Ray prompt not configured.")
    prompt = prompt_template.replace("{user_input}", user_input[:1000])
    try:
        queries = call_llm_json(prompt, model_name)
    except Exception as e:
        raise HTTPException(500, f"LLM error: {e}")
    if not isinstance(queries, list):
        queries = [queries]
    result = []
    for q in queries:
        if isinstance(q, dict) and "query" in q:
            result.append({
                "platform": str(q.get("platform", "Search")),
                "description": str(q.get("description", "")),
                "query": str(q.get("query", "")),
            })
    return result


# ── GitHub Miner ──────────────────────────────────────────────

_GH_API = "https://api.github.com"


def _gh_get(endpoint: str, token: str) -> dict | list | None:
    """Authenticated GitHub API GET request."""
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github.v3+json"}
    try:
        r = httpx.get(f"{_GH_API}{endpoint}", headers=headers, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None


@app.post("/github/mine")
async def github_mine(request: Request):
    """Search GitHub repos by keywords, extract contributors, return candidate profiles."""
    _auth.require_auth(request)
    body = await request.json()
    keywords = (body.get("keywords") or "").strip()
    location = (body.get("location") or "").strip().lower()
    min_stars = int(body.get("min_stars") or 100)
    if len(keywords) < 3:
        raise HTTPException(400, "Keywords must be at least 3 characters.")
    config = _core.load_config()
    token = _resolve_github_token(config)
    if not token:
        raise HTTPException(400, "GitHub PAT not configured. Set it at /setup or via GITHUB_TOKEN env var.")

    # Phase 1: search repos
    import urllib.parse
    query = urllib.parse.quote(keywords)
    repos_data = _gh_get(f"/search/repositories?q={query}+stars:>={min_stars}&sort=stars&order=desc&per_page=10", token)
    if not repos_data or not repos_data.get("items"):
        return []

    # Phase 2: extract contributors
    unique_users = {}
    for repo in repos_data["items"]:
        repo_name = repo["full_name"]
        contribs = _gh_get(f"/repos/{repo_name}/contributors?per_page=5", token)
        if not contribs:
            continue
        for c in contribs:
            login = c.get("login", "")
            if login and login not in unique_users and "[bot]" not in login:
                unique_users[login] = {"repo": repo_name, "contributions": c.get("contributions", 0)}

    # Phase 3: fetch profiles and filter by location
    candidates = []
    for login, meta in unique_users.items():
        profile = _gh_get(f"/users/{login}", token)
        if not profile:
            continue
        user_loc = profile.get("location") or ""
        if location and location not in user_loc.lower():
            continue
        candidates.append({
            "login": login,
            "name": profile.get("name") or login,
            "location": user_loc or "Unknown",
            "company": profile.get("company") or "Independent",
            "email": profile.get("email") or "Hidden",
            "bio": profile.get("bio") or "",
            "html_url": profile.get("html_url", ""),
            "repo": meta["repo"],
            "contributions": meta["contributions"],
        })

    return candidates


@app.post("/github/import")
async def github_import(request: Request):
    """Import a GitHub profile: fetch repos, LLM→CV JSON, save to store."""
    _auth.require_auth(request)
    body = await request.json()
    login = (body.get("login") or "").strip()
    if not login:
        raise HTTPException(400, "GitHub login required.")
    config = _core.load_config()
    token = _resolve_github_token(config)
    if not token:
        raise HTTPException(400, "GitHub PAT not configured. Set it at /setup or via GITHUB_TOKEN env var.")
    model_name = choose_model_name(config)
    api_key = resolve_api_key(APP_DIR, config)
    configure_gemini(api_key)

    # Fetch user profile and repos
    user_data = _gh_get(f"/users/{login}", token)
    if not user_data:
        raise HTTPException(502, f"Failed to fetch GitHub profile for {login}")
    repos_data = _gh_get(f"/users/{login}/repos?sort=updated&per_page=10", token) or []

    # Build prompt
    prompt_template = config.get("prompt_github", _core.DEFAULT_PROMPTS.get("prompt_github", ""))
    gh_full_data = json.dumps({"user": user_data, "recent_repos": repos_data}, ensure_ascii=False)
    prompt = prompt_template.replace("{prompt_schema_only}", _core.CV_JSON_SCHEMA).replace("{gh_full_data}", gh_full_data)

    # Call LLM
    _JOB_SEMAPHORE.acquire()
    try:
        cv_json = call_llm_json(prompt, model_name)
    except Exception as e:
        raise HTTPException(500, f"LLM error: {e}")
    finally:
        _JOB_SEMAPHORE.release()

    if not isinstance(cv_json, dict):
        raise HTTPException(500, "LLM returned invalid CV data")

    # Sanitize
    cv_json = _core.sanitize_json(cv_json)

    # Generate store ID from login
    store_id = hashlib.sha256(f"github_{login}".encode()).hexdigest()

    # Save to store (dedup by name)
    source_filename = f"github.com/{login}"
    name = cv_json.get("basics", {}).get("name", login)
    existing = _find_store_by_name(name)
    if existing:
        store_id = existing.stem
    _save_to_store(store_id, cv_json, source_filename)

    return {
        "ok": True,
        "store_id": store_id,
        "name": name,
    }
